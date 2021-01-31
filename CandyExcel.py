#! /usr/bin/env python3

import openpyxl
from openpyxl.reader.excel import load_workbook
import pandas as pd
import requests
from itertools import islice
from functools import reduce
from bs4 import BeautifulSoup as bs
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment

class CandyExcel:
	''' Excel data class '''

	ND = "not defined"

	def __init__(self):
		self._raw_data = None
		self._new_data = pd.DataFrame()
		self._out_wb = None
		self._out_ws = None

	def _get_data(self, page):
		''' Downloads a page data from the website 'page'
			and returns a parser object '''
		try:
			if (not page):
				return (None)
			if (page[:8] != "https://"):
				return (page)
			headers = {"User-agent":"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.80 Safari/537.36"}
			res = requests.get(page, headers=headers, verify=False, timeout = 1.5)
			soup = bs(res.text, "html.parser")
			return (soup)
		except:
			return (self.ND)

	def _iget_bakerstore(self, parser):
		''' Gets price from the bakerstore page's parser object '''
		if (not parser):
			return ("")
		if (type(parser) == str):
			return (self.ND)
		raw = parser.find("span", {"class" : "autocalc-product-special"})
		try:
			if (raw):
				price = raw.string
			else:
				price = parser.find("span", {"class" : "autocalc-product-price"}).string
			return int(price.replace(".0", ""))
		except:
			return (self.ND)

	def _iget_tortomaster(self, parser):
		''' Gets price from the tortomaster page's parser object '''
		if (not parser):
			return ("")
		if (type(parser) == str):
			return (self.ND)
		try:
			price = "".join(parser.find_all("span", {"class" : "price"})[0].text[:-1].split(" "))
			return int(price)
		except:
			return (self.ND)

	def _iget_vtk(self, parser):
		''' Gets price from the vtk page's parser object '''
		if (not parser):
			return ("")
		if (type(parser) == str):
			return (self.ND)
		try:
			price = parser.find("span", {"class" : "tprice-value"}).string
			return int(price.replace(" ", ""))
		except:
			return (self.ND)

	def set_raw_data(self, path, GUIParams):
		wb = openpyxl.load_workbook(path)
		ws = wb.active

		data = ws.values
		cols = next(data)[1:]
		data = list(data)
		idx = [r[0] for r in data]
		data = (islice(r, 1, None) for r in data)
		self._raw_data = pd.DataFrame(data, index=idx, columns=cols)
		wb.close()
		GUIParams["pb"]["value"] += 5

	def create_new_data(self, GUIParams):
		GUIParams["lbl"].config(text = "Загружаем с VTK...")
		self._new_data["VTK"]			= self._raw_data["VTK"].apply(lambda x: self._iget_vtk(parser = self._get_data(x)) if x != None else None)
		GUIParams["pb"]["value"] += 20
		GUIParams["lbl"].config(text = "Загружаем с bakerstore...")
		self._new_data["bakerstore"]	= self._raw_data["bakerstore"].apply(lambda x: self._iget_bakerstore(parser = self._get_data(x)) if x != None else None)
		GUIParams["pb"]["value"] += 20
		GUIParams["lbl"].config(text = "Загружаем с tortomaster...")
		self._new_data["tortomaster"]	= self._raw_data["tortomaster"].apply(lambda x: self._iget_tortomaster(parser = self._get_data(x)) if x != None else None)
		GUIParams["pb"]["value"] += 20

	def save_data(self, GUIParams):
		print(self._new_data)
		GUIParams["lbl"].config(text = "Сохраняем файл...")
		empty = pd.DataFrame()
		empty[""] = ""
		dfs = [self._new_data, empty, self._raw_data["VTK"], self._raw_data["bakerstore"], self._raw_data["tortomaster"]]
		for i in range(5):
			dfs[i] = dfs[i].reset_index(drop = True)
		for d in dfs:
			print(d)
		self._new_data = pd.concat(dfs, axis = 1, ignore_index = True)
		print(self._new_data)

		self._out_wb = openpyxl.Workbook()
		self._out_ws = self._out_wb.active

		inds = self._raw_data.index.tolist()
		cols = self._raw_data.columns.values.tolist()
		cols.append(None)
		cols += self._raw_data.columns.values.tolist()
		values = self._new_data.values.tolist()

		cols.insert(0, None)
		print(cols)
		for i in range(len(inds)):
			print(inds[i], values[i])
		self._out_ws.append(cols)
		for i in range(len(values)):
			values[i].insert(0, inds[i])
			self._out_ws.append(values[i])

		GUIParams["pb"]["value"] += 15

	def prettify_data(self, GUIParams):
		GUIParams["lbl"].config(text = "Фоорматируем данные...")
		for cell in self._out_ws['A'] + self._out_ws[1]:
			if (cell.value):
				cell.alignment = Alignment(horizontal='left')

		for cell in self._out_ws[1]:
			if (cell.value):
				cell.font = Font(bold = True)

		for row in range(1, self._out_ws.max_row + 1):
			if self._out_ws.cell(row = row, column = 1).value and not self._out_ws.cell(row = row, column = 2).value:
				self._out_ws.cell(row = row, column = 1).font = Font(bold = True)
	
		for cell in self._out_ws['B'] + self._out_ws['C'] + self._out_ws['D']:
			cell.number_format = "General"

		for row in range(2, self._new_data.shape[0] + 2):
			m = 1000000000
			min_col = 2
			for col in [2,3,4]:
				if self._out_ws.cell(row = row, column = col).value and \
					type(self._out_ws.cell(row = row, column = col).value) == int and \
					int(self._out_ws.cell(row = row, column = col).value) < int(m):
					m = int(self._out_ws.cell(row = row, column = col).value)
					min_col = col
			if self._out_ws.cell(row = row, column = min_col).value and \
			type(self._out_ws.cell(row = row, column = min_col).value) == int:
				self._out_ws.cell(row = row, column = min_col).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type = 'solid')

		self._out_ws.column_dimensions['A'].width = 80
		self._out_ws.column_dimensions['B'].width = 20
		self._out_ws.column_dimensions['C'].width = 20
		self._out_ws.column_dimensions['D'].width = 20

		GUIParams["pb"]["value"] += 10
	'''
	/ --- Deprecated --- /
	/ ------------------ /
	/ ------------------ /
	/ ------------------ /
	/ ---   Actual   --- /
	'''
	def	clone_update(self, filename):
		self._out_wb = load_workbook(filename)
		self._out_ws = self._out_wb.active

		for r in self._out_ws.iter_rows():
			for c in r:
				if (c.value and type(c.value) == str and c.value[:8] == "https://"):
					if ("vtk" in str(c.value)):
						c.value = self._iget_vtk(parser = self._get_data(c.value))
					if ("bakerstore" in str(c.value)):
						c.value = self._iget_bakerstore(parser = self._get_data(c.value))
					if ("tortomaster" in str(c.value)):
						c.value = self._iget_tortomaster(parser = self._get_data(c.value))
			if (type(r[1].value) == int or type(r[2].value) == int or type(r[3].value) == int):
				m = min(r[1:], key = lambda x: x.value if type(x.value) == int else 1000000000)
				m.fill = PatternFill(start_color='92D050', end_color='92D0FF', fill_type = 'solid')

	def close_data(self, filename):
		self._out_wb.save("./" + filename.split('/')[-1].split('.')[-2] + "_out.xlsx")
		self._out_wb.close()
