#! /usr/bin/env python3

from openpyxl.reader.excel import load_workbook
import pandas as pd
import requests
from itertools import islice
from functools import reduce
from bs4 import BeautifulSoup as bs
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from ntpath import basename
import openpyxl
from string import Template
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders

class CandyExcel:
	''' Excel data class '''

	ND = "not defined"

	def __init__(self):
		self._raw_data = None
		self._new_data = pd.DataFrame()
		self._out_wb = None
		self._out_ws = None

	def _get_data(self, page):
		''' Downloads a page data from the website 'page'
			and returns a parser object '''
		try:
			if (not page):
				return (None)
			if (page[:8] != "https://"):
				return (page)
			headers = {"User-agent":"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/47.0.2526.80 Safari/537.36"}
			res = requests.get(page, headers=headers, verify=False, timeout = 3)
			soup = bs(res.text, "html.parser")
			return (soup)
		except:
			return (self.ND)

	def _iget_bakerstore(self, parser):
		''' Gets price from the bakerstore page's parser object '''
		if (not parser):
			return ("")
		if (type(parser) == str):
			return (self.ND)
		raw = parser.find("span", {"class" : "autocalc-product-special"})
		try:
			if (raw):
				price = raw.string
			else:
				price = parser.find("span", {"class" : "autocalc-product-price"}).string
			return int(price.replace(".0", ""))
		except:
			return (self.ND)

	def _iget_tortomaster(self, parser):
		''' Gets price from the tortomaster page's parser object '''
		if (not parser):
			return ("")
		if (type(parser) == str):
			return (self.ND)
		try:
			price = "".join(parser.find_all("span", {"class" : "price"})[0].text[:-1].split(" "))
			return int(price)
		except:
			return (self.ND)

	def _iget_vtk(self, parser):
		''' Gets price from the vtk page's parser object '''
		if (not parser):
			return ("")
		if (type(parser) == str):
			return (self.ND)
		try:
			price = parser.find("span", {"class" : "tprice-value"}).string
			return int(price.replace(" ", ""))
		except:
			return (self.ND)

	def prettify(self):
		for r in self._out_ws.iter_rows():
			if (r[0].value and (type(r[1].value) == int or type(r[2].value) == int or type(r[3].value) == int)):
				m = min(r[1:], key = lambda x: x.value if type(x.value) == int else 1000000000)
				m.fill = PatternFill(start_color='DDFBD8', end_color='DDFBD8', fill_type = 'solid')

	def	clone_update(self, filename):
		self._out_wb = load_workbook(filename)
		self._out_ws = self._out_wb.active
		self._out_ws.title = "Прайс-лист"

		for r in self._out_ws.iter_rows():
			if (r[0].value):
				for c in r:
					link = c.value
					if (c.value and type(c.value) == str and c.value[:8] == "https://"):
						if ("vtk" in str(c.value)):
							c.value = self._iget_vtk(parser = self._get_data(c.value))
							c.hyperlink = link
						if ("bakerstore" in str(c.value)):
							c.value = self._iget_bakerstore(parser = self._get_data(c.value))
							c.hyperlink = link
						if ("tortomaster" in str(c.value)):
							c.value = self._iget_tortomaster(parser = self._get_data(c.value))
							c.hyperlink = link

	def calc_changes(self, filename):

		old = openpyxl.load_workbook(basename(filename).split('.')[0] + "_out.xlsx").active
		new = self._out_ws

		changes = self._out_wb.copy_worksheet(self._out_ws)
		changes.title = "Изменения"

		for ro, rn, rc in zip(
			old.iter_rows(),
			new.iter_rows(),
			changes.iter_rows()
		):
			for co, cn, cc in zip(ro, rn, rc):
				if (co.value and cn.value and type(co.value) == type(cn.value) == int):
					if (co.value == cn.value):
						cc.value = 0
					else:
						cc.value = cn.value - co.value
						cc.number_format = "+0;-0"
						if cc.value > 0:
							cc.fill = PatternFill(start_color='F8D8D6', end_color='F8D8D6', fill_type = 'solid')
						if cc.value < 0:
							cc.fill = PatternFill(start_color='DDFBD8', end_color='DDFBD8', fill_type = 'solid')
				else:
					cc.value = cn.value

	def close_data(self, filename):
		self._out_wb.save("./" + filename.split('/')[-1].split('.')[-2] + "_out.xlsx")
		self._out_wb.close()

def mailing(filename, attachment_f):

	data = pd.read_excel(filename)
	names = data[data.columns[0]].to_list()
	emails = data[data.columns[1]].to_list()

	template = Template(
	"Уважаемый ${NAME},\n" + \
	"это рассылка от Клубничкиной!\n" + \
	"В прикрепленный файлах вы найдете всю информацию о ценах на самые необходимые кондитерские ингридиенты!\n" + \
	"С любовью,\n" + \
	"ваша Клубничкина!")

	s = smtplib.SMTP(host='smtp.gmail.com', port=587)
	s.starttls()
	s.login("feedflax83@gmail.com", "CHEBurashkaKAKA278")
	
	fromaddr = "feedflax83@gmail.com"
	for name, email in zip(names, emails):
		toaddr = email
		msg = MIMEMultipart()
		msg['From'] = fromaddr
		msg['To'] = toaddr
		msg['Subject'] = "Рассылка от Клубничкиной"
		body = template.substitute(NAME = name)
		msg.attach(MIMEText(body, 'plain'))
		attachment = open(attachment_f, "rb")
		p = MIMEBase('application', 'octet-stream')
		p.set_payload((attachment).read())
		encoders.encode_base64(p)
		p.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
		msg.attach(p)
		text = msg.as_string()
		s.sendmail(fromaddr, toaddr, text)
		del msg
	s.quit()